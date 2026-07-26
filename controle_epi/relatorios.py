"""
relatorios.py
Geração de relatórios de entregas de EPI em Excel (.xlsx) e PDF.
"""

from io import BytesIO
from datetime import datetime

import database as db

# ---------------------------------------------------------------------------
# EXCEL
# ---------------------------------------------------------------------------

STATUS_COR_HEX = {
    "Válido": "C6EFCE",              # verde claro
    "Próximo do vencimento": "FFEB9C",  # amarelo claro
    "Vencido": "FFC7CE",              # vermelho claro
    "Devolvido": "D9D9D9",            # cinza claro
}


def gerar_excel(entregas) -> bytes:
    """Gera um relatório em Excel (.xlsx) a partir de uma lista de entregas
    (linhas retornadas por database.get_entregas). Retorna os bytes do arquivo."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Entregas de EPI"

    colunas = [
        "Funcionário", "Setor", "EPI", "Data da Entrega",
        "Data de Validade", "Status", "Data de Devolução",
    ]
    ws.append(colunas)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E5395", end_color="2E5395", fill_type="solid")
    for col_idx, _ in enumerate(colunas, start=1):
        celula = ws.cell(row=1, column=col_idx)
        celula.font = header_font
        celula.fill = header_fill
        celula.alignment = Alignment(horizontal="center")

    for entrega in entregas:
        status = db.status_entrega(entrega["data_validade"], entrega["devolvido"])
        linha = [
            entrega["funcionario"],
            entrega["setor"] or "-",
            entrega["epi"],
            entrega["data_entrega"],
            entrega["data_validade"],
            status,
            entrega["data_devolucao"] or "-",
        ]
        ws.append(linha)

        cor_hex = STATUS_COR_HEX.get(status)
        if cor_hex:
            linha_idx = ws.max_row
            fill = PatternFill(start_color=cor_hex, end_color=cor_hex, fill_type="solid")
            ws.cell(row=linha_idx, column=6).fill = fill  # coluna "Status"

    # Ajusta largura das colunas automaticamente (aproximado)
    for col_idx, titulo in enumerate(colunas, start=1):
        max_len = len(titulo)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            valor = row[0].value
            if valor:
                max_len = max(max_len, len(str(valor)))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 4

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def gerar_pdf(entregas) -> bytes:
    """Gera um relatório em PDF a partir de uma lista de entregas.
    Retorna os bytes do arquivo."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    )
    from reportlab.lib.styles import getSampleStyleSheet

    STATUS_COR_PDF = {
        "Válido": colors.HexColor("#C6EFCE"),
        "Próximo do vencimento": colors.HexColor("#FFEB9C"),
        "Vencido": colors.HexColor("#FFC7CE"),
        "Devolvido": colors.HexColor("#D9D9D9"),
    }

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1.5 * cm, rightMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    styles = getSampleStyleSheet()
    elementos = []

    elementos.append(Paragraph("🦺 Relatório de Controle de EPI", styles["Title"]))
    data_geracao = datetime.now().strftime("%d/%m/%Y %H:%M")
    elementos.append(Paragraph(f"Gerado em: {data_geracao}", styles["Normal"]))
    elementos.append(Spacer(1, 0.5 * cm))

    total = len(entregas)
    vencidos = sum(1 for e in entregas if db.status_entrega(e["data_validade"], e["devolvido"]) == "Vencido")
    proximos = sum(
        1 for e in entregas if db.status_entrega(e["data_validade"], e["devolvido"]) == "Próximo do vencimento"
    )
    resumo = f"Total de entregas: {total} | Vencidos: {vencidos} | Próximos do vencimento: {proximos}"
    elementos.append(Paragraph(resumo, styles["Normal"]))
    elementos.append(Spacer(1, 0.5 * cm))

    cabecalho = ["Funcionário", "Setor", "EPI", "Entrega", "Validade", "Status", "Devolução"]
    dados_tabela = [cabecalho]
    cores_linha = []

    for entrega in entregas:
        status = db.status_entrega(entrega["data_validade"], entrega["devolvido"])
        dados_tabela.append([
            entrega["funcionario"],
            entrega["setor"] or "-",
            entrega["epi"],
            entrega["data_entrega"],
            entrega["data_validade"],
            status,
            entrega["data_devolucao"] or "-",
        ])
        cores_linha.append(STATUS_COR_PDF.get(status, colors.white))

    tabela = Table(dados_tabela, repeatRows=1)
    estilo = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E5395")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]
    for idx, cor in enumerate(cores_linha, start=1):
        estilo.append(("BACKGROUND", (5, idx), (5, idx), cor))  # colore só a coluna Status

    tabela.setStyle(TableStyle(estilo))
    elementos.append(tabela)

    doc.build(elementos)
    buffer.seek(0)
    return buffer.getvalue()
